# Imports
import os
import re
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import accuracy_score,classification_report
from imblearn.over_sampling import SMOTE  # For oversampling the minority class
import tkinter as tk
from tkinter import Tk, filedialog
from sklearn.base import BaseEstimator
from matplotlib.colors import ListedColormap
import shutil
import openpyxl
from openpyxl import load_workbook

def screen_data(df, features, target): # data screening function
    """
    Creates a subset of the desired features and applies listwise deletion (deletes rows if any feature is missing).
    
    Parameters:
        df (DataFrame): Original DataFrame.
        features (list): List of feature column names to retain.
        target (str): optional one off parameter.
    
    Returns:
        subset (DataFrame): Processed DataFrame after listwise deletion.
    """
    if target is not None:
        subset = df[features + [target]].dropna(axis=0, how='any')
    else:
        subset = df[features].dropna(axis=0, how='any')
        
    
    if subset.empty:
        print(f"Dataset is empty.")
        return None,

    return subset

def select_output_directory():# Get the output directory from user input
    """
    This function prompts the user to select an output directory using a graphical file dialog.

    Returns:
        str: The path of the selected output directory.

    Raises:
        Exception: If no directory is selected by the user.
    """
    # Initialize the Tkinter root window
    root = Tk()

    # Lift the root window to the top
    root.lift()
    root.attributes('-topmost', True)
    
    # Open a directory selection dialog and store the selected directory
    directory = filedialog.askdirectory(title="Select Output Directory", parent=root)

    # Withdraw the root window after the dialog is closed
    root.withdraw()

    # Check if a directory was selected
    if directory:
        return directory
    else:
        raise Exception("No directory selected.")  

def save_location_pre_processing(kernel, output_path=None):
    """
    Prepares the directory and file path for saving SVM classification reports.
    
    Parameters:
    output_path (str): The desired output directory path. If None, the user will be prompted to select one.
    kernel (str): The kernel type used for the SVM.
    
    Returns:
    tuple: A tuple containing the output directory path and the full file path for excel classification reports.
    """
    
    folder_name = f"SVM_{kernel}_results"
    file_name = f"SVM_{kernel}_classification_report.xlsx"

    # If no output path is provided, prompt the user to select a directory
    if output_path is None:
        output_path = select_output_directory()
        
    output_directory = os.path.join(output_path, folder_name)
    output_file = os.path.join(output_directory, file_name)
  
    # Delete the directory if it already exists 
    if os.path.exists(output_directory):
        shutil.rmtree(output_directory)
        
    # Check if the file already exists, and remove if it does
    if os.path.isfile(output_file):
        os.remove(output_file)
    
    return output_directory, output_file

def save_to_excel(output_file, sheet_name, results_df, description):
    """
    Saves a DataFrame to an existing Excel workbook, with the ability to add a description.
    
    Parameters:
    output_file (str): The path to the Excel file where the DataFrame will be saved.
    sheet_name (str): The name of the sheet where the DataFrame will be written.
    results_df (pd.DataFrame): The DataFrame containing results to be saved.
    description (str): A description to be placed at the beginning of the sheet.
    """

    # Use Pandas Excel writer with Openpyxl as the engine
    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        # Write the DataFrame to the specified sheet, starting at row 2
        results_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=True)
        
        # Write the description in the first cell of the sheet
        worksheet = writer.sheets[sheet_name]
        worksheet.cell(row=1, column=1, value=description)

def get_marker_map(subset, target):
    """
    Generates a mapping from unique '3-tier rank' values to specific marker symbols.

    Parameters:
    subset (pandas.DataFrame): The data subset containing the '3-tier rank' values.
    target (str): The column name in 'subset' which contains the '3-tier rank' values.

    Returns:
    tuple:
        - marker_map (dict): A dictionary mapping each unique '3-tier rank' value to a marker symbol.
        - unique_rates (List of strings): A list of the sorted unique '3-tier rank' values.

    Raises:
    AssertionError: If the number of unique '3-tier rank' values is not exactly three. 
    
    Note: could modify this by adding more markers so the same tests can be done for the 5 tier rank
    """

    # Retrieve and sort the unique '3-tier rank' values from the target column
    unique_rates = subset[target].unique()
    unique_rates.sort()

    # Ensure there are exactly three unique '3-tier rank' values
    assert len(unique_rates) == 3, "The dataset must have exactly 3 unique '3-tier rank' values"

    # Define a list of distinct markers corresponding to the three unique '3-tier rank' values
    markers = ['^', 'o', 's']
    marker_map = {rate: markers[i] for i, rate in enumerate(unique_rates)}

    return marker_map, unique_rates

def plot_decision_boundary(df_train, df_test, target, feature_x, 
                            feature_y, marker_map, rates, labels, kernel_choice,
                            model, train_accuracy, test_accuracy, 
                            best_param, plot_save_name):

    # Define boundaries for the plot
    x_min, x_max = np.min([df_train[feature_x].min(), df_test[feature_x].min()]) - 1, np.max([df_train[feature_x].max(), df_test[feature_x].max()]) + 1
    y_min, y_max = np.min([df_train[feature_y].min(), df_test[feature_y].min()]) - 1, np.max([df_train[feature_y].max(), df_test[feature_y].max()]) + 1
    xx, yy = np.meshgrid(np.arange(x_min, x_max, 0.01), np.arange(y_min, y_max, 0.01))

    # Predict on grid points
    Z = model.predict(np.c_[xx.ravel(), yy.ravel()])
    # Z = label_encoder.inverse_transform(Z)
    Z = Z.reshape(xx.shape)
  
    # Create a color map using a sequential or diverging colormap
    colors = plt.cm.coolwarm(np.linspace(0, 1, len(rates)))
    cmap = ListedColormap(colors)
  
    # Set up the plots
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))

    # Plot the decision boundaries on both subplots    
    contour1 = axes[0].contourf(xx, yy, Z,levels = 3, alpha=0.5, cmap=cmap) 
    contour2 = axes[1].contourf(xx, yy, Z,levels = 3, alpha=0.5, cmap=cmap)
    

    # Plot the training and test data with different markers for each class
    for i, rate in enumerate(rates):
        train_mask = df_train[target] == rate
        test_mask = df_test[target] == rate
        axes[0].scatter(df_train[feature_x][train_mask], df_train[feature_y][train_mask], 
                        marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)
        axes[1].scatter(df_test[feature_x][test_mask], df_test[feature_y][test_mask], 
                        marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)
        # decision boundary and prediction check 
        # pred_mask = df_test["prediction"] == rate
        # axes[0].scatter(df_test[feature_x][pred_mask], df_test[feature_y][pred_mask], 
        #                 marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)

    if kernel_choice == 'poly':     
        main_title = f"Decision Boundary and Ground Truth: {feature_x} vs {feature_y}\nKernel: Degree {best_param} polynomial\nNote: features are standardized"
        fig.suptitle(main_title)
        axes[0].set_title(f"Training Set\nAccuracy: {train_accuracy:.2f}")
        axes[0].set_xlabel(feature_x)
        axes[0].set_ylabel(feature_y)
        axes[0].legend(title=target, labels = rates)
        
        axes[1].set_title(f"Testing Set\nAccuracy: {test_accuracy:.2f}")
        axes[1].set_xlabel(feature_x)
        axes[1].set_ylabel(feature_y)
        axes[1].legend(title=target, labels = rates)
    else:
        
        main_title = f"Decision Boundary and Ground Truth: {feature_x} vs {feature_y}\nKernel: {kernel_choice} with regularization hyper-parameter C: {best_param}\nNote: features are standardized"
        fig.suptitle(main_title)
        axes[0].set_title(f"Training Set\nAccuracy: {train_accuracy:.2f}")
        axes[0].set_xlabel(feature_x)
        axes[0].set_ylabel(feature_y)
        axes[0].legend(title=target, labels = rates)
        
        axes[1].set_title(f"Testing Set\nAccuracy: {test_accuracy:.2f}")
        axes[1].set_xlabel(feature_x)
        axes[1].set_ylabel(feature_y)
        axes[1].legend(title=target, labels = rates)

    # Adjust layout
    plt.tight_layout()

    # Save the plot
    plt.savefig(plot_save_name)
    # plt.show()

def evaluate_feature_pairs(df, features, target, kernel_choice, output_path, output_file, save_name_mapping, param_grid, n_splits=3):
    """
    Evaluates all unique pairs of features using Support Vector Machine (SVM) with specified kernel and hyperparameters.
    
    Parameters:
    df (pandas.DataFrame): The DataFrame containing the data.
    features (list of str): List of feature column names to be evaluated in pairs.
    target (str): The target column name containing '3-tier rank' values.
    kernel_choice (str): The choice of kernel for SVM ('linear', 'rbf', etc.).
    output_path (str): Directory path to save the plots.
    output_file (str): File path to save the Excel outputs.
    save_name_mapping (dict): Mapping from feature names to desired save names for files.
    n_splits (int): Number of splits for k-fold cross-validation (default is 3).
    param_grid (dict): Dictionary specifying the hyperparameters for GridSearchCV (default includes 'C' with values [0.1, 1, 10]).
    
    Returns:
    list: A list containing tuples with feature pairs and their evaluation metrics.
    """

    results = []

    # Get marker mapping and labels
    marker_map, rates = get_marker_map(df, target)
   
    # Encode rates to integer labels
    label_encoder = LabelEncoder()
    labels = label_encoder.fit_transform(rates)
    
    # Transform the target values from strings to int using label encoder  
    y = label_encoder.transform(df[target].values)

    # Handle class imbalance using SMOTE
    smote = SMOTE(random_state=42)
    X_resampled, y_resampled = smote.fit_resample(df[features].values, y)

    # Standardize features for SVM
    scaler = StandardScaler()
    X_resampled = scaler.fit_transform(X_resampled)
    
    # Set up k-fold cross-validation
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    
    # Split data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X_resampled, y_resampled, test_size=0.3, random_state=42)
    
    
    
    for i in range(len(features)):
        for j in range(i + 1, len(features)):
            feature_x = features[i]
            feature_y = features[j]
            
            # Extract feature pairs
            X_train_features = X_train[:, [i, j]]
            X_test_features = X_test[:, [i, j]]
            
            # Create save names
            save_name_x = save_name_mapping[feature_x]
            save_name_y = save_name_mapping[feature_y]
            plot_save_name = os.path.join(output_path, f"SVM_{kernel_choice}_plot_{save_name_x}_vs_{save_name_y}.png")
            sheet_name = f"{save_name_x} and {save_name_y}"
    
        
            # Process ground truth for plotting using resampled and standardized data
            # Prepare the training data
            y_train_rates = label_encoder.inverse_transform(y_train)
            df_train = pd.DataFrame(X_train_features, columns=[feature_x, feature_y])
            df_train[target] = y_train_rates
            df_train["encoded rates"] = y_train
            
            # Prepare the testing data
            y_test_rates = label_encoder.inverse_transform(y_test)
            df_test = pd.DataFrame(X_test_features, columns=[feature_x, feature_y])
            df_test[target] = y_test_rates
            df_test["encoded rates"] = y_test
            
                        
            # Initialize the SVM model
            if kernel_choice == "poly":
                svc = SVC(kernel='poly', gamma='scale')
                grid_search = GridSearchCV(svc, param_grid, cv=kf)
                
            elif kernel_choice == 'rbf':
                svc = SVC(kernel=kernel_choice, gamma='scale')
                # Use GridSearchCV with k-fold cross-validation
                grid_search = GridSearchCV(svc, param_grid, cv=kf)
            else:
                svc = SVC(kernel=kernel_choice)
                grid_search = GridSearchCV(svc, param_grid, cv=kf)
            
            # Use GridSearchCV with k-fold cross-validation
            grid_search.fit(X_train_features, y_train)
            
            # Retrieve the best model and its scores
            best_model = grid_search.best_estimator_
            best_params = grid_search.best_params_
            
            # Calculate training and test accuracy
            train_accuracy = best_model.score(X_train_features, y_train)
            test_accuracy = best_model.score(X_test_features, y_test)
                        
            # Predict the degradation rate using the test data and best model
            y_test_pred = best_model.predict(X_test_features)
            y_test_pred_rates = label_encoder.inverse_transform(y_test_pred)
            df_test["prediction"] = y_test_pred_rates
            
            if kernel_choice == "poly":
                # Save results
                results.append((feature_x, feature_y, train_accuracy, test_accuracy, grid_search.best_score_, best_model, best_params['degree']))

                # Description for excel sheet
                description = (f"Best model report for SVM model with {kernel_choice} kernel using degree = {best_params['degree']}, and using features: {feature_x} and "
                            f"{feature_y}")
                            
                # Print results
                print(f"SVM Model: {feature_x} vs {feature_y} with {kernel_choice} kernel")
                print(f"Best degree model: {best_params['degree']}")
                print(f"Training Accuracy: {train_accuracy:.2f}")
                print(f"Testing Accuracy: {test_accuracy:.2f}")
                print("\nClassification Report (Test Data):")
                print(classification_report(y_test, y_test_pred, target_names=label_encoder.classes_, zero_division=0))
                # Plot decision boundary
                plot_decision_boundary(df_train, df_test, target, feature_x, feature_y, marker_map, rates, labels, 
                                kernel_choice, best_model, train_accuracy, test_accuracy, best_params['degree'], plot_save_name)
            else:
                # Save results
                results.append((feature_x, feature_y, train_accuracy, test_accuracy, grid_search.best_score_, best_model, best_params['C']))
            
                # Description for excel sheet
                description = (f"Best model report for SVM model with {kernel_choice} kernel using features: {feature_x} and "
                            f"{feature_y} with C hyperparameter = {best_params['C']}")
                
                # Print results
                print(f"SVM Model: {feature_x} vs {feature_y} with {kernel_choice} kernel")
                print(f"Best C hyperparameter: {best_params['C']}")
                print(f"Training Accuracy: {train_accuracy:.2f}")
                print(f"Testing Accuracy: {test_accuracy:.2f}")
                print("\nClassification Report (Test Data):")
                print(classification_report(y_test, y_test_pred, target_names=label_encoder.classes_, zero_division=0))       
                    
                   
                
                # Plot decision boundary
                plot_decision_boundary(df_train, df_test, target, feature_x, feature_y, marker_map, rates, labels, 
                                kernel_choice, best_model, train_accuracy, test_accuracy, best_params['C'], plot_save_name)
                
            # Convert the classification report dictionary to a pandas DataFrame for saving to Excel
            report = classification_report(y_test, y_test_pred, target_names=label_encoder.classes_, zero_division=0, output_dict=True)
            report.update({"accuracy": {"precision": None, "recall": None, "f1-score": report["accuracy"], "support": report['macro avg']['support']}})
            report_df = pd.DataFrame(report).transpose()
            report_df = report_df.round(2)
        
            # save results in excel
            save_to_excel(output_file, sheet_name, report_df, description)
                
                
    return results

def on_button_click(root, button_id):
    global kernel 
    kernel = button_id
    # root.withdraw()
    # root.quit()
    root.destroy()
    
def choose_kernel():
    # Initialize the Tkinter root window
    root = Tk()

    # Lift the root window to the top
    root.lift()
    root.attributes('-topmost', True)
    
    # Create the main window
    root.title("Select the kernel")
    
    # Create three buttons and assign them to the window
    button1 = tk.Button(root, text="linear", command=lambda: on_button_click(root,"linear"))
    button1.pack()

    button2 = tk.Button(root, text="rbf", command=lambda: on_button_click(root,"rbf"))
    button2.pack()

    button3 = tk.Button(root, text="polynomial", command=lambda: on_button_click(root,"poly"))
    button3.pack()
    root.mainloop()





# data set loading, feature definitions,alpha for significance test

# Load the Excel file
df = pd.read_excel('Project Dataset.xlsx', sheet_name='data', engine='openpyxl')

# renaming features for clarity
df = df.rename(columns={'BOD (% day-1)': 'BOD (% /day)',
                        'wt. loss (% day-1)':'wt. loss (% /day)',                        
                        'den (g mL-1)':'den (g/mL)',
                        "% cryst": "% crystallinity", 
                        "enthalpy (J g-1)": "enthalpy (J/g)", 
                        "LogP(SA)-1 (Å-2)": "LogP/(SA)", 
                        "Mw (kg mol-1)": "Mw (kg/mol)",
                        "Mn (kg mol-1)": "Mn (kg/mol)",
                        '3-tier rank': 'Degradation Rate'
                        })

# grouping feature for plots
target = 'Degradation Rate'

# Select the features and target
features = ["% crystallinity", 
            "enthalpy (J/g)", 
            "Tg (°C)", 
            "LogP/(SA)", 
            "Mw (kg/mol)"
            ]

target = 'Degradation Rate'

# cant use special characters in the save files so now it maps to an acceptable name
save_name_mapping = {'den (g/mL)':'density',
                    'total sp3 C':'total sp3 C', 
                    '% sp3 C':'percent sp3 C', 
                    'LogP/(SA)': 'LogPSA',
                    'Mw (kg/mol)': 'MW', 
                    'Mn (kg/mol)': 'Mn', 
                    'Mw/Mn':'Ratio MwMn',
                    'Tg (°C)':'Tg',
                    'Tm (°C)':'Tm',
                    '% crystallinity':'percent cryst',
                    'enthalpy (J/g)':'enthalpy',
                    '(Tw-Tg)/(LogP/SA)':'TWTW_LogPSA'
                    }
# C hyperparameters tested 
C = {'C': [0.1, 1, 5, 10, 20, 50, 75, 100, 200, 500]}
# Degree hyperparameters tested 
degree = {'degree': [2,3,4,3,5,6,7,8]}
from tkinter import *

kernel = None
# Initialize output path as None
output_path = None

def main():
    choose_kernel()
    print("Note that the data preprocessing includes standard scaling and resampling.(standardScaler() and SMOTE fit_transform())")

    # Get output path from user input and pre process the directory
    output_path, output_file = save_location_pre_processing(kernel)

    # create an output folder 
    os.makedirs(output_path)
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Save the workbook to the specified file
    wb.save(output_file)   


    # Perform data screening
    data = screen_data(df, features, target)

    # Evaluate feature pairs and visualize
    if kernel == "poly":
        results = evaluate_feature_pairs(data, features, target, kernel, output_path, output_file, save_name_mapping, param_grid = degree)
        # Display results
        print("Summary Evaluation Results:")
        print("Degree tested: \n ", degree )
        print("=" * 40)
        for feature_x, feature_y, train_acc, test_acc, best_score, best_model, best_degree in results:
            print(f"Features: {feature_x} vs {feature_y}")
            print(f"Best C: {best_degree:.2f}")
            print(f"Training Accuracy: {train_acc:.2f}")
            print(f"Testing Accuracy: {test_acc:.2f}")
            print("-" * 40)

    else:
        results = evaluate_feature_pairs(data, features, target, kernel, output_path, output_file, save_name_mapping, param_grid = C)
        # Display results
        print("Summary Evaluation Results:")
        print("C hyperparameters tested: \n ", C )
        print("=" * 40)
        for feature_x, feature_y, train_acc, test_acc, best_score, best_model, best_C in results:
            print(f"Features: {feature_x} vs {feature_y}")
            print(f"Best C: {best_C:.2f}")
            print(f"Training Accuracy: {train_acc:.2f}")
            print(f"Testing Accuracy: {test_acc:.2f}")
            print("-" * 40)
        
    # Post processing workbook
    # Load the workbook
    wb = load_workbook(output_file)
    # Remove the first sheet
    wb.remove(wb[wb.sheetnames[0]])
    # Save the workbook to apply changes
    wb.save(output_file)
        
    return 0
    
    
if __name__ == "__main__":
    main()