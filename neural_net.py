import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
import matplotlib.pyplot as plt
# Imports
import os
import re
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split, KFold, GridSearchCV
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import accuracy_score,classification_report
from imblearn.over_sampling import SMOTE  # For oversampling the minority class
import tkinter as tk
from tkinter import Tk, filedialog
from tkinter import *
from sklearn.base import BaseEstimator
from matplotlib.colors import ListedColormap
import shutil
import openpyxl
from openpyxl import load_workbook

# Define the neural network class
class ShallowNeuralNetwork(nn.Module):
    def __init__(self, input_num, hidden_num, output_num):
        super(ShallowNeuralNetwork, self).__init__()
        self.hidden = nn.Linear(input_num, hidden_num) # Hidden layer
        self.output = nn.Linear(hidden_num, output_num) # Output layer
        self.relu = nn.ReLU() # ReLU activation
        
    def forward(self, x):
        x = self.relu(self.hidden(x))
        out = self.output(x)
        return out

    def predict(self, x):
        with torch.no_grad():
            logits = self.forward(x)
            predictions = torch.argmax(logits, dim=1)
        return predictions
    
def screen_data(df, features, target): # data screening function
    """
    Creates a subset of the desired features and applies listwise deletion (deletes rows if any feature is missing).
    
    Parameters:
        df (DataFrame): Original DataFrame.
        features (list): List of feature column names to retain.
        target (str): optional one off parameter.
    
    Returns:
        subset (DataFrame): Processed DataFrame after listwise deletion.
    """
    if target is not None:
        subset = df[features + [target]].dropna(axis=0, how='any')
    else:
        subset = df[features].dropna(axis=0, how='any')
        
    
    if subset.empty:
        print(f"Dataset is empty.")
        return None,

    return subset

def select_output_directory():# Get the output directory from user input
    """
    This function prompts the user to select an output directory using a graphical file dialog.

    Returns:
        str: The path of the selected output directory.

    Raises:
        Exception: If no directory is selected by the user.
    """
    # Initialize the Tkinter root window
    root = Tk()

    # Lift the root window to the top
    root.lift()
    root.attributes('-topmost', True)
    
    # Open a directory selection dialog and store the selected directory
    directory = filedialog.askdirectory(title="Select Output Directory", parent=root)

    # Withdraw the root window after the dialog is closed
    root.withdraw()

    # Check if a directory was selected
    if directory:
        return directory
    else:
        raise Exception("No directory selected.")  

def save_location_pre_processing(output_path=None):
    """
    Prepares the directory and file path for saving SVM classification reports.
    
    Parameters:
    output_path (str): The desired output directory path. If None, the user will be prompted to select one.
    kernel (str): The kernel type used for the SVM.
    
    Returns:
    tuple: A tuple containing the output directory path and the full file path for excel classification reports.
    """
    
    folder_name = f"neural_net"
    file_name = f"neural_net_report.xlsx"

    # If no output path is provided, prompt the user to select a directory
    if output_path is None:
        output_path = select_output_directory()
        
    output_directory = os.path.join(output_path, folder_name)
    output_file = os.path.join(output_directory, file_name)
  
    # Delete the directory if it already exists 
    if os.path.exists(output_directory):
        shutil.rmtree(output_directory)
        
    # Check if the file already exists, and remove if it does
    if os.path.isfile(output_file):
        os.remove(output_file)
    
    return output_directory, output_file

def save_to_excel(output_file, sheet_name, results_df, grid_search_results, description1,description2):
    """
    This function saves given DataFrames to an existing Excel workbook, creating separate sheets for detailed results
    and grid search outcomes. It also allows including descriptions at the top of each sheet to provide context
    or additional information related to the results.

    Parameters:
        - output_file (str): The path to the Excel file where the DataFrames will be saved. This file should exist beforehand.
        - sheet_name (str): The name of the sheet where the main results DataFrame will be written.
        - results_df (pd.DataFrame): The DataFrame containing the main evaluation results to be saved.
        - grid_search_results (pd.DataFrame): The DataFrame containing detailed grid search results to be saved on a
        separate sheet named `sheet_name_2`.
        - description1 (str): A description to be written at the top of the `sheet_name` sheet.
        - description2 (str): A description to be written at the top of the `sheet_name_2` sheet (the grid search results sheet).

    Functionality:
        - Opens the specified Excel file in append mode and writes the provided DataFrames to specific sheets.
        - Adds contextual descriptions at the top of each sheet before writing the DataFrame content.
        - If the specified sheets already exist in the workbook, their contents will be replaced.

    Note: 
        - This function uses the 'openpyxl' engine, which supports Excel file manipulation in Python.
        - Ensure that `output_file` exists and is writable; otherwise, `pd.ExcelWriter` may throw an error.
    """
    sheet_name2 = f"{sheet_name}_2"
    # Use Pandas Excel writer with Openpyxl as the engine
    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        # Write the DataFrame to the specified sheet, starting at row 2
        results_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=True)
        
        # Write the description in the first cell of the sheet
        worksheet = writer.sheets[sheet_name]
        worksheet.cell(row=1, column=1, value=description1)
        
        # Write the DataFrame to the specified sheet, starting at row 2
        grid_search_results.to_excel(writer, sheet_name=sheet_name2, startrow=2, index=True)
        
        # Write the description in the first cell of the sheet
        worksheet = writer.sheets[sheet_name2]
        worksheet.cell(row=1, column=1, value=description2)

def get_marker_map(subset, target):
    """
    Generates a mapping from unique '3-tier rank' values to specific marker symbols.

    Parameters:
    subset (pandas.DataFrame): The data subset containing the '3-tier rank' values.
    target (str): The column name in 'subset' which contains the '3-tier rank' values.

    Returns:
    tuple:
        - marker_map (dict): A dictionary mapping each unique '3-tier rank' value to a marker symbol.
        - unique_rates (List of strings): A list of the sorted unique '3-tier rank' values.

    Raises:
    AssertionError: If the number of unique '3-tier rank' values is not exactly three. 
    
    Note: could modify this by adding more markers so the same tests can be done for the 5 tier rank
    """

    # Retrieve and sort the unique '3-tier rank' values from the target column
    unique_rates = subset[target].unique()
    unique_rates.sort()

    # Ensure there are exactly three unique '3-tier rank' values
    assert len(unique_rates) == 3, "The dataset must have exactly 3 unique '3-tier rank' values"

    # Define a list of distinct markers corresponding to the three unique '3-tier rank' values
    markers = ['^', 'o', 's']
    marker_map = {rate: markers[i] for i, rate in enumerate(unique_rates)}

    return marker_map, unique_rates

def on_button_click(root, button_id):
    global kernel 
    kernel = button_id
    root.destroy()
   
def choose_kernel():
    # Initialize the Tkinter root window
    root = Tk()

    # Lift the root window to the top
    root.lift()
    root.attributes('-topmost', True)
    
    # Create the main window
    root.title("Select the kernel")
    
    # Create three buttons and assign them to the window
    button1 = tk.Button(root, text="linear", command=lambda: on_button_click(root,"linear"))
    button1.pack()

    button2 = tk.Button(root, text="rbf", command=lambda: on_button_click(root,"rbf"))
    button2.pack()

    button3 = tk.Button(root, text="polynomial", command=lambda: on_button_click(root,"poly"))
    button3.pack()
    root.mainloop()

def plot_decision_boundary(df_train, df_test, target, feature_x, 
                            feature_y, marker_map, rates, labels, kernel_choice,
                            model, train_accuracy, test_accuracy, 
                            best_params, plot_save_name):
    '''
    This function plots the decision boundaries of a classification model alongside the actual ground truth data points
    for both training and testing datasets. It supports plotting over a two-dimensional feature space defined by 'feature_x'
    and 'feature_y'. The function uses contour plots to illustrate the decision regions and scatter plots to show the data points.
    Different markers are used for different classes based on the provided `marker_map`.

    Parameters:
    - df_train (pd.DataFrame): Training data containing features and target variable.
    - df_test (pd.DataFrame): Testing data containing features and target variable.
    - target (str): The name of the target variable/column.
    - feature_x (str): The name of the first feature for the x-axis.
    - feature_y (str): The name of the second feature for the y-axis.
    - marker_map (dict): Dictionary mapping each class label to a specific marker style.
    - rates (list): List of unique class labels/rates in the data.
    - labels (list): Descriptive labels for each class/rate to be displayed in legend.
    - kernel_choice (str): Type of kernel used in the model (e.g., 'poly', 'rbf', 'linear').
    - model: Trained classification model that supports `.predict()` method.
    - train_accuracy (float): Training accuracy of the model.
    - test_accuracy (float): Testing accuracy of the model.
    - best_params (dict): Dictionary containing the best hyperparameters for the model.
    - plot_save_name (str): The file name/path to save the plot image.

    Functionality:
    - Computes the decision boundary on a specified mesh grid.
    - Visualizes the decision boundary and classifies regions using the model's predictions.
    - Displays training and test data points with specific markers per class.
    - Titles the plots with model and dataset details, including accuracy and kernel hyperparameters.
    - Saves the generated plot to the specified file.
'''

    # Define boundaries for the plot
    x_min, x_max = np.min([df_train[feature_x].min(), df_test[feature_x].min()]) - 1, np.max([df_train[feature_x].max(), df_test[feature_x].max()]) + 1
    y_min, y_max = np.min([df_train[feature_y].min(), df_test[feature_y].min()]) - 1, np.max([df_train[feature_y].max(), df_test[feature_y].max()]) + 1
    xx, yy = np.meshgrid(np.arange(x_min, x_max, 0.01), np.arange(y_min, y_max, 0.01))

    # Predict on grid points
    Z = model.predict(np.c_[xx.ravel(), yy.ravel()])
    # Z = label_encoder.inverse_transform(Z)
    Z = Z.reshape(xx.shape)
  
    # Create a color map using a sequential or diverging colormap
    colors = plt.cm.coolwarm(np.linspace(0, 1, len(rates)))
    cmap = ListedColormap(colors)
  
    # Set up the plots
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))

    # Plot the decision boundaries on both subplots    
    contour1 = axes[0].contourf(xx, yy, Z,levels = 3, alpha=0.5, cmap=cmap) 
    contour2 = axes[1].contourf(xx, yy, Z,levels = 3, alpha=0.5, cmap=cmap)
    

    # Plot the training and test data with different markers for each class
    for i, rate in enumerate(rates):
        train_mask = df_train[target] == rate
        test_mask = df_test[target] == rate
        axes[0].scatter(df_train[feature_x][train_mask], df_train[feature_y][train_mask], 
                        marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)
        axes[1].scatter(df_test[feature_x][test_mask], df_test[feature_y][test_mask], 
                        marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)
        # decision boundary and prediction check 
        # pred_mask = df_test["prediction"] == rate
        # axes[0].scatter(df_test[feature_x][pred_mask], df_test[feature_y][pred_mask], 
        #                 marker=marker_map[rate], color=colors[i], edgecolors='k', label=labels[i], alpha=0.7)

    if kernel_choice == 'poly':     
        main_title = f"Decision Boundary and Ground Truth: {feature_x} vs {feature_y}\nKernel: polynomial\nHyperparameters: Degree = {best_params['degree']}, Gamma = {best_params['gamma']} \nNote: features are standardized"
        
    elif kernel_choice == "rbf":
        main_title = f"Decision Boundary and Ground Truth: {feature_x} vs {feature_y}\nKernel: {kernel_choice}\n Hyperparameters: C = {best_params['C']}, Gamma = {best_params['gamma']}\nNote: features are standardized"

    elif kernel_choice == "linear":
        main_title = f"Decision Boundary and Ground Truth: {feature_x} vs {feature_y}\nKernel: {kernel_choice}\n Hyperparameters: C = {best_params['C']}\nNote: features are standardized"

    # Plot parameters
    fig.suptitle(main_title)
    axes[0].set_title(f"Training Set\nAccuracy: {train_accuracy:.2f}")
    axes[0].set_xlabel(feature_x)
    axes[0].set_ylabel(feature_y)
    axes[0].legend(title=target, labels = rates)
    
    axes[1].set_title(f"Testing Set\nAccuracy: {test_accuracy:.2f}")
    axes[1].set_xlabel(feature_x)
    axes[1].set_ylabel(feature_y)
    axes[1].legend(title=target, labels = rates)
    
    # Adjust layout
    plt.tight_layout()

    # Save the plot
    plt.savefig(plot_save_name)
    # plt.show()

def evaluate_feature_pairs(df, features, target, kernel_choice, output_path, output_file, save_name_mapping, param_grid, res_features, n_splits=3):
    """
    Evaluate Feature Pairs using SVM

    This function evaluates all unique pairs of features from a given DataFrame using a Support Vector Machine (SVM) 
    with specified kernel and hyperparameters. It performs the following tasks for each pair of features:

    1. Handles class imbalance using the Synthetic Minority Over-sampling Technique (SMOTE).
    2. Standardizes the features for SVM compatibility.
    3. Splits the data for training and testing.
    4. Utilizes GridSearchCV to find the best hyperparameters for the given kernel.
    5. Trains the best SVM model and evaluates its performance on the test dataset.
    6. Plots decision boundaries and saves the plots.
    7. Saves evaluation results and detailed grid search results to an Excel file for further analysis.

    Parameters:
        - df (pandas.DataFrame): The DataFrame containing the input data.
        - features (list of str): List of feature column names to be evaluated in pairs.
        - target (str): The target column name containing class labels.
        - kernel_choice (str): The choice of kernel for the SVM ('linear', 'rbf', 'poly', etc.).
        - output_path (str): Directory path to save the decision boundary plots.
        - output_file (str): File path to save the evaluation results in Excel format.
        - save_name_mapping (dict): Dictionary mapping feature names to customized filenames for saving plots.
        - n_splits (int): Number of splits for k-fold cross-validation. Default is 3.
        - param_grid (dict): Dictionary specifying the hyperparameters for grid search using GridSearchCV.
        - res_features (list of str): List of GridSearchCV result features to be included in the results Excel file.

    Returns:
        - list: Returns a list of tuples where each tuple contains:
            * feature_x (str): First feature in the evaluated pair.
            * feature_y (str): Second feature in the evaluated pair.
            * train_accuracy (float): Training accuracy of the best model.
            * test_accuracy (float): Testing accuracy of the best model.
            * best_cv_score (float): Best cross-validation score obtained during grid search.
            * best_model: The best SVM model obtained from grid search.
            * best_params (dict): Best hyperparameters obtained from grid search.

    The function saves decision boundary plots and Excel files containing detailed SVM evaluation reports.
    """

    results = []

    # Get marker mapping and labels
    marker_map, rates = get_marker_map(df, target)
   
    # Encode rates to integer labels
    label_encoder = LabelEncoder()
    labels = label_encoder.fit_transform(rates)
    
    # Transform the target values from strings to int using label encoder  
    y = label_encoder.transform(df[target].values)

    # Handle class imbalance using SMOTE
    smote = SMOTE(random_state=42)
    X_resampled, y_resampled = smote.fit_resample(df[features].values, y)

    # Standardize features for SVM
    scaler = StandardScaler()
    X_resampled = scaler.fit_transform(X_resampled)
    
    # Set up k-fold cross-validation
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    
    # Split data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X_resampled, y_resampled, test_size=0.3, random_state=42)
    
    
    
    for i in range(len(features)):
        for j in range(i + 1, len(features)): 
            feature_x = features[i]
            feature_y = features[j]
            
            # Extract feature pairs
            X_train_features = X_train[:, [i, j]]
            X_test_features = X_test[:, [i, j]]
            
            # Create save names
            save_name_x = save_name_mapping[feature_x]
            save_name_y = save_name_mapping[feature_y]
            plot_save_name = os.path.join(output_path, f"SVM_{kernel_choice}_plot_{save_name_x}_vs_{save_name_y}.png")
            sheet_name = f"{save_name_x} and {save_name_y}"
    
        
            # Process ground truth for plotting using resampled and standardized data
            # Prepare the training data
            y_train_rates = label_encoder.inverse_transform(y_train)
            df_train = pd.DataFrame(X_train_features, columns=[feature_x, feature_y])
            df_train[target] = y_train_rates
            df_train["encoded rates"] = y_train
            
            # Prepare the testing data
            y_test_rates = label_encoder.inverse_transform(y_test)
            df_test = pd.DataFrame(X_test_features, columns=[feature_x, feature_y])
            df_test[target] = y_test_rates
            df_test["encoded rates"] = y_test
            
                        
            # Initialize the SVM model
            svc = SVC(kernel= kernel_choice)
            grid_search = GridSearchCV(svc, param_grid, cv=kf)
            
            # Use GridSearchCV with k-fold cross-validation
            grid_search.fit(X_train_features, y_train)
            
            # Retrieve the best model and its scores
            best_model = grid_search.best_estimator_
            best_params = grid_search.best_params_
            
            # Calculate training and test accuracy
            train_accuracy = best_model.score(X_train_features, y_train)
            test_accuracy = best_model.score(X_test_features, y_test)
                        
            # Predict the degradation rate using the test data and best model
            y_test_pred = best_model.predict(X_test_features)
            y_test_pred_rates = label_encoder.inverse_transform(y_test_pred)
            df_test["prediction"] = y_test_pred_rates
            
            # saving results to return for summary
            results.append((feature_x, feature_y, train_accuracy, test_accuracy, grid_search.best_score_, best_model, best_params))
            
            # Plot decision boundary
            plot_decision_boundary(df_train, df_test, target, feature_x, feature_y, marker_map, rates, labels, 
                                    kernel_choice, best_model, train_accuracy, test_accuracy, best_params, plot_save_name)
            
            
            # Retrieve entire results of grid search, screen for specific features and sort by score
            res = grid_search.cv_results_
            res_df = pd.DataFrame(res)
            res_df = res_df[res_features]
            res_df.sort_values("rank_test_score", ascending=True, inplace=True, kind = 'mergesort')
            
            if kernel_choice == "poly":
                # Description for excel sheet
                description1 = (f"Best model report for SVM model with {kernel_choice} kernel. Hyperparameters: Degree = {best_params['degree']}, Gamma = {best_params['gamma']}. Features: {feature_x} and "
                            f"{feature_y}")
            
            elif kernel_choice == "rbf":
                # Description for excel sheet
                description1 = (f"Best model report for SVM model with {kernel_choice} kernel. Hyperparameters: C = {best_params['C']}, Gamma = {best_params['gamma']}. Features: {feature_x} and "
                            f"{feature_y}")  
                
            elif kernel_choice == "linear":            
                # Description for excel sheet
                description1 = (f"Best model report for SVM model with {kernel_choice} kernel. Hyperparameters: C = {best_params['C']}. Features: {feature_x} and "
                            f"{feature_y}")  
            # description for excel sheet 2
            description2 = (f"Grid search results for SVM model with {kernel_choice} kernel. Features: {feature_x} and "
                            f"{feature_y}")  
                
                
            # Convert the classification report dictionary to a pandas DataFrame for saving to Excel
            report = classification_report(y_test, y_test_pred, target_names=label_encoder.classes_, zero_division=0, output_dict=True)
            report.update({"accuracy": {"precision": None, "recall": None, "f1-score": report["accuracy"], "support": report['macro avg']['support']}})
            report_df = pd.DataFrame(report).transpose()
            report_df = report_df.round(2)
        
            # save results in excel
            save_to_excel(output_file, sheet_name, report_df, res_df, description1, description2)
                 
                
    return results

# data set loading, feature definitions,alpha for significance test
# Load the Excel file
df = pd.read_excel('Project Dataset.xlsx', sheet_name='data', engine='openpyxl')

# renaming features for clarity
df = df.rename(columns={'BOD (% day-1)': 'BOD (% /day)',
                        'wt. loss (% day-1)':'wt. loss (% /day)',                        
                        'den (g mL-1)':'den (g/mL)',
                        "% cryst": "% crystallinity", 
                        "enthalpy (J g-1)": "enthalpy (J/g)", 
                        "LogP(SA)-1 (Å-2)": "LogP/(SA)", 
                        "Mw (kg mol-1)": "Mw (kg/mol)",
                        "Mn (kg mol-1)": "Mn (kg/mol)",
                        '3-tier rank': 'Degradation Rate'
                        })

# grouping feature for plots
target = 'Degradation Rate'

# Select the features and target
features = [
            "Tg (°C)", 
            "LogP/(SA)", 
            ]

target = 'Degradation Rate'

# Set device to CPU
device = torch.device("cpu")

def main():
    # # Get output path from user input and pre process the directory
    # output_path, output_file = save_location_pre_processing()

    # # create an output folder 
    # os.makedirs(output_path)
    # # Create a new workbook
    # wb = openpyxl.Workbook()
    # # Save the workbook to the specified file
    # wb.save(output_file)   

    # Perform data screening
    data = screen_data(df, features, target)
    
    # Get marker mapping and labels
    marker_map, rates = get_marker_map(data, target)
    
    # Encode rates to integer labels
    label_encoder = LabelEncoder()
    labels = label_encoder.fit_transform(rates)
    
    # Transform the target values from strings to int using label encoder  
    y = label_encoder.transform(data[target].values)
   
    # Feature-target split
    X = data[features].values.astype(np.float32)
       
    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, shuffle=True, random_state=42)
    
    # convert to tensors
    X_train = torch.tensor(X_train, dtype=torch.float32).to(device)
    X_test = torch.tensor(X_test, dtype=torch.float32).to(device)
    y_train = torch.tensor(y_train, dtype=torch.long).to(device)
    y_test = torch.tensor(y_test, dtype=torch.long).to(device)
    
    # Define model parameters
    input_num = X_train.shape[1] # Number of features
    hidden_num = 2 # Hidden layer size
    output_num = len(np.unique(y_train)) # Number of classes

    # Define learning rates to test
    learning_rates = [0.001, 0.005, 0.01, 0.02, 0.05, 0.1]
    train_accuracies = []
    test_accuracies = []
    # Train and evaluate model for each learning rate
    for lr in learning_rates:
        model = ShallowNeuralNetwork(input_num, hidden_num, output_num).to(device)
        criterion = nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=lr)
        # Training loop
        num_epochs = 500
        for epoch in range(num_epochs):
            model.train()
            y_pred = model(X_train)
            loss = criterion(y_pred, y_train)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
        # Evaluate the model
        model.eval()
        with torch.no_grad():
            train_preds = model.predict(X_train)
            test_preds = model.predict(X_test)
        # Calculate accuracies
        train_accuracy = accuracy_score(y_train.cpu().numpy(), train_preds.cpu().numpy())
        test_accuracy = accuracy_score(y_test.cpu().numpy(), test_preds.cpu().numpy())
        train_accuracies.append(train_accuracy)
        test_accuracies.append(test_accuracy)
        print(f"LR: {lr}, Train Accuracy: {train_accuracy:.4f}, Test Accuracy:{test_accuracy:.4f}")
        
    # Plot accuracy vs. learning rate
    plt.figure(figsize=(10, 6))
    plt.plot(learning_rates, train_accuracies, label="Train Accuracy", marker='o')
    plt.plot(learning_rates, test_accuracies, label="Test Accuracy", marker='o')
    plt.xscale('log')
    plt.xlabel("Learning Rate (log scale)")
    plt.ylabel("Accuracy")
    plt.title("Accuracy vs. Learning Rate")
    plt.legend()
    plt.grid(True)
    plt.show()
    return 0

    
if __name__ == "__main__":
    main()